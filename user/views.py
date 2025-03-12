from django.shortcuts import render, redirect, get_object_or_404
from .models import Question, Student, Choice
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from .forms import StudentRegistrationForm
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
import time
from django.contrib import messages
from django.utils.translation import gettext as _
from django.utils import timezone
from django.utils.translation import activate
from docx import Document
from django.db.models import Q
from openpyxl import Workbook
import openpyxl
from django.utils.timezone import now

def register_student(request):
    if request.method == "POST":
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            first_name = form.cleaned_data["first_name"]
            password = form.cleaned_data["password"]
            student = form.save(commit=False)
            student.save()
            unique_username = student.unique_code
            user = User.objects.create_user(
                username=unique_username,
                password=password,
                first_name=first_name
            )
            student.user = user
            student.save()
            request.session['expected_unique_code'] = student.unique_code
            request.session['user_id'] = user.id
            # Сбрасываем таймер для нового пользователя
            request.session[f'timer_start_{user.id}'] = int(time.time() * 1000)  # Время в миллисекундах
            return render(request, "verify_unique_code.html", {"form": form})
    else:
        form = StudentRegistrationForm()
    return render(request, "register.html", {"form": form})

def verify_unique_code(request):
    if request.method == "POST":
        entered_code = request.POST.get("unique_code")
        expected_code = request.session.get("expected_unique_code")
        user_id = request.session.get("user_id")
        language = request.POST.get("language")
        difficulty = request.POST.get("difficulty")

        if not all([entered_code, expected_code, user_id, language, difficulty]):
            return render(request, "verify_unique_code.html", {
                "error": "Заполните все поля: код, язык и сложность."
            })

        if entered_code == expected_code:
            user = User.objects.get(id=user_id)
            student = user.student
            student.login_time = now()
            student.save()
            login(request, user)

            # 1. Находим все модули, которые соответствуют языку и сложности
            modules = Question.objects.filter(
                Q(module__contains=difficulty) & Q(module__contains=language)
            ).values('module').distinct()

            # Проверяем, что модулей достаточно (например, 5)
            if len(modules) < 5:
                return render(request, "verify_unique_code.html", {
                    "error": f"Недостаточно модулей для {difficulty} на языке {language}. Найдено {len(modules)} модулей, нужно 5."
                })

            # 2. Для каждого модуля выбираем по 10 случайных вопросов
            selected_questions = []
            for module in modules:
                module_name = module['module']  # Например, "module_1_easy_uz"
                questions_in_module = list(Question.objects.filter(module=module_name).order_by('?')[:10])

                # Проверяем, что в модуле достаточно вопросов
                if len(questions_in_module) < 10:
                    return render(request, "verify_unique_code.html", {
                        "error": f"В модуле {module_name} недостаточно вопросов: найдено {len(questions_in_module)}, нужно минимум 10."
                    })

                selected_questions.extend(questions_in_module)

            # 3. Проверяем общее количество вопросов (должно быть 50)
            if len(selected_questions) != 50:
                return render(request, "verify_unique_code.html", {
                    "error": f"Общее количество вопросов не равно 50: найдено {len(selected_questions)}."
                })

            # 4. Сохраняем выбранные вопросы и параметры в сессии
            request.session['quiz_questions'] = [q.id for q in selected_questions]
            request.session['language'] = language
            request.session['difficulty'] = difficulty

            # Очищаем сессионные данные верификации
            if 'expected_unique_code' in request.session:
                del request.session['expected_unique_code']
            if 'user_id' in request.session:
                del request.session['user_id']

            return redirect("home")
        else:
            return render(request, "verify_unique_code.html", {
                "error": "Неверный уникальный код. Попробуйте снова."
            })
    return redirect("register")

def student_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            student = user.student  # Получаем связанный объект Student
            student.login_time = now()  # Логируем время входа
            student.save()
            return redirect("home")
        else:
            # Добавляем сообщение об ошибке через messages
            messages.error(request, "Неверные данные. Проверьте имя пользователя и пароль.")
            return render(request, "login.html")
    return render(request, "login.html")

def student_logout(request):
    logout(request)
    return redirect('login')
#
# def dashboard(request):
#     return render(request, 'dashboard.html', {'user': request.user})

@login_required
def quiz_list(request):
    # Получаем язык из сессии и активируем его
    language = request.session.get('language', 'uz')  # По умолчанию 'uz'
    activate(language)  # Активируем язык для текущего запроса

    quiz_questions_ids = request.session.get('quiz_questions', [])
    if not quiz_questions_ids:
        return redirect("verify_unique_code")

    questions = Question.objects.filter(id__in=quiz_questions_ids).prefetch_related('choices')
    total_questions = len(questions)
    question_list = [{'id': q.id, 'text': q.text} for q in questions]

    current_question = request.GET.get('q', '1')
    try:
        current_question = int(current_question)
    except ValueError:
        current_question = 1
    current_question = max(1, min(current_question, total_questions))

    question = questions[current_question - 1] if total_questions > 0 else None
    selected_answer = request.session.get('answers', {}).get(f'answer_{question.id}') if question else None

    user = request.user
    student, created = Student.objects.get_or_create(user=user)

    if not student.test_start_time and not student.test_end_time:
        student.test_start_time = now()
        student.save()
        request.session[f'timer_start_{user.id}'] = int(time.time() * 1000)
        request.session.modified = True

    if 'answers' not in request.session:
        request.session['answers'] = {}

    initial_time = 3600
    if 'timer_reset' in request.GET:
        request.session['answers'] = {}
        if f'timer_start_{user.id}' in request.session:
            del request.session[f'timer_start_{user.id}']
        request.session.modified = True

    if request.method == 'POST' and 'submit_quiz' in request.POST:
        submitted_answers = request.session.get('answers', {}).copy()
        correct_count = 0
        user_answers = []

        for q in questions:
            q_id = f'answer_{q.id}'
            answer = submitted_answers.get(q_id, None)
            choice = q.choices.first()
            if choice:
                correct_answer = choice.is_correct.lower()
                normalized_answer = answer.lower() if answer else None
                is_correct = normalized_answer == correct_answer if answer else False

                user_answers.append({
                    'question_text': q.text,
                    'user_answer': normalized_answer.upper() if normalized_answer else _("Не отвечено"),
                    'correct_answer': correct_answer.upper(),
                    'is_correct': is_correct
                })
                if is_correct:
                    correct_count += 1

        score = (correct_count / total_questions) * 50 if total_questions > 0 else 0
        student.score = score
        student.test_end_time = now()
        student.save()
        request.session['answers'] = {}
        if f'timer_start_{user.id}' in request.session:
            del request.session[f'timer_start_{user.id}']
        if 'quiz_questions' in request.session:
            del request.session['quiz_questions']
        if 'language' in request.session:
            del request.session['language']
        if 'difficulty' in request.session:
            del request.session['difficulty']
        request.session.modified = True

        return render(request, 'results.html', {
            'score': score,
            'total_questions': total_questions,
            'total_score': score * 2,
            'current_language': language,
            'user': user,
            'user_answers': user_answers
        })

    return render(request, 'index.html', {
        'question': question,
        'current_language': language,
        'user': user,
        'current_question': current_question,
        'total_questions': total_questions,
        'initial_time': initial_time,
        'answers': request.session.get('answers', {}),
        'selected_answer': selected_answer,
        'question_list': question_list,
        'start_time': request.session.get(f'timer_start_{user.id}')
    })

@login_required
def set_timer(request):
    if request.method == 'POST':
        start_time = request.POST.get('start_time')
        if start_time:
            request.session[f'timer_start_{request.user.id}'] = int(start_time)
            request.session.modified = True
            return JsonResponse({'status': 'success'})
        return JsonResponse({'status': 'error'}, status=400)
    return JsonResponse({'status': 'method not allowed'}, status=405)


MONTHS = {
    'January': 'Январь',
    'February': 'Февраль',
    'March': 'Март',
    'April': 'Апрель',
    'May': 'Май',
    'June': 'Июнь',
    'July': 'Июль',
    'August': 'Август',
    'September': 'Сентябрь',
    'October': 'Октябрь',
    'November': 'Ноябрь',
    'December': 'Декабрь'
}

@login_required
def download_student_report(request):
    ids_str = request.GET.get('ids', '')
    student_ids = [int(id) for id in ids_str.split(',') if id.isdigit()]

    if not student_ids:
        return HttpResponse("Не выбрано ни одного студента.", status=400)

    students = Student.objects.filter(id__in=student_ids)

    # Создаем Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Студенты"

    # Заголовки таблицы
    headers = ['Фамилия', 'Имя', 'Уникальный код', 'Баллы', 'Время входа', 'Время начала теста', 'Время окончания теста']
    for col_num, header in enumerate(headers, 1):
        ws[f'{chr(64 + col_num)}1'] = header
        ws[f'{chr(64 + col_num)}1'].font = openpyxl.styles.Font(bold=True)

    # Заполняем данными студентов
    for row_num, student in enumerate(students, 2):
        ws[f'A{row_num}'] = student.last_name
        ws[f'B{row_num}'] = student.first_name
        ws[f'C{row_num}'] = student.unique_code
        ws[f'D{row_num}'] = student.score
        if student.login_time:
            # Преобразуем дату в локальный часовой пояс (Asia/Tashkent)
            local_time = timezone.localtime(student.login_time)
            # Форматируем дату с заменой месяца на русский
            date_str = local_time.strftime("%d %B %Y г. %H:%M")
            month = local_time.strftime("%B")
            date_str = date_str.replace(month, MONTHS[month])
            ws[f'E{row_num}'] = date_str
        else:
            ws[f'E{row_num}'] = 'Не указано'
        if student.test_start_time:
            # Преобразуем дату в локальный часовой пояс (Asia/Tashkent)
            local_time = timezone.localtime(student.test_start_time)
            # Форматируем дату с заменой месяца на русский
            date_str = local_time.strftime("%d %B %Y г. %H:%M")
            month = local_time.strftime("%B")
            date_str = date_str.replace(month, MONTHS[month])
            ws[f'F{row_num}'] = date_str
        else:
            ws[f'F{row_num}'] = 'Не указано'
        if student.test_end_time:
            # Преобразуем дату в локальный часовой пояс (Asia/Tashkent)
            local_time = timezone.localtime(student.test_end_time)
            # Форматируем дату с заменой месяца на русский
            date_str = local_time.strftime("%d %B %Y г. %H:%M")
            month = local_time.strftime("%B")
            date_str = date_str.replace(month, MONTHS[month])
            ws[f'G{row_num}'] = date_str
        else:
            ws[f'G{row_num}'] = 'Не указано'

    # Автоматическая настройка ширины колонок
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Настраиваем ответ для скачивания
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students_report.xlsx'
    wb.save(response)
    return response

@require_POST
@login_required
def save_answer(request):
    question_id = request.POST.get('question_id')
    answer = request.POST.get('answer')
    if question_id and answer and question_id.isdigit():
        if 'answers' not in request.session:
            request.session['answers'] = {}

        # Сохраняем ответ в сессии
        request.session['answers'][f'answer_{question_id}'] = answer
        request.session.modified = True
        print(f"Saved answer for question {question_id}: {answer}")  # Отладка
        print(f"Session after save: {request.session.get('answers')}")  # Дополнительная отладка

        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

